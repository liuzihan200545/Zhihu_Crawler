from selenium import webdriver
import openpyxl
import streamlit as st
import pandas as pd

file_name = 'example1.xlsx'
wb = openpyxl.load_workbook(file_name)
ws = wb["Sheet"]

driver = webdriver.Edge()

def function(n):
    excel_name = 'B' + str(n+2)
    url = ws[excel_name].value
    driver.get("https:" + url)

def search(name):
    for i in range(1,ws.max_row+1):
        if(name in ws['A'+str(i)].value):
            #print(ws['A'+str(i)].value)
            st.info(ws['A'+str(i)].value+str(i-2))


def get_data_from_excel():
  df = pd.read_excel(io="./example1.xlsx",engine="openpyxl",sheet_name="Sheet")
  return df

df = get_data_from_excel()
st.info("知乎收藏夹工具")
st.write(df)
with st.form('input_form'):
    text = st.text_area('请输入要打开的网页的编号:', '')
    if(st.form_submit_button('Submit')):
        function(int(text))

with st.form('input_form1'):
    text = st.text_area('请输入要查询的内容:', '')
    if(st.form_submit_button('Submit')):
        search(text)








